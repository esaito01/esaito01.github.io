import openpyxl
import os

wb = openpyxl.Workbook()
print(wb)

print(wb.get_sheet_names())
sheetname = wb.get_sheet_names()
#sheet = wb.get_sheet_by_name(sheetname)

sheet = wb.get_sheet_by_name('Sheet')

print(sheet.cell(row=1, column = 1))

#sheet.cell(row = 1, column = 1) = 'Hello'
sheet["A2"] = "Hello2"
os.chdir('C:\\Temporary\\Excel')

wb.save('example2.xlsx')

sheet2 = wb.create_sheet()
print(wb.get_sheet_names())
print(sheet2.title)
sheet2.title = "My New Sheet name"
print(wb.get_sheet_names())


wb.save('example2.xlsx')

wb.close()